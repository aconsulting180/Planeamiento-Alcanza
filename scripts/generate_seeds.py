import openpyxl
import json

wb = openpyxl.load_workbook('Plan de Gestión 2026 - Entregable final.xlsx', data_only=True)
ws = wb['PTOTAL']

# Extract data grouped by area
areas = {}  # area_name -> list of rows

for r in range(5, ws.max_row + 1):
    grupo = str(ws.cell(r, 2).value or '').strip()
    og_full = str(ws.cell(r, 3).value or '').strip()
    oe_full = str(ws.cell(r, 4).value or '').strip()
    kpi = str(ws.cell(r, 5).value or '').strip()
    cod = str(ws.cell(r, 6).value or '').strip().replace('\n','').strip()
    act = str(ws.cell(r, 7).value or '').strip()
    ent = str(ws.cell(r, 8).value or '').strip()
    resp = str(ws.cell(r, 9).value or '').strip()
    
    months_active = []
    for c in range(10, 22):  # Ene=10 to Dic=21
        v = ws.cell(r, c).value
        if v and str(v).strip():
            months_active.append(c - 9)  # 1=Ene, 12=Dic
    
    if not grupo or grupo == 'Grupo':
        continue
    
    # Skip rows that are sub-headers (no code, act looks like a subtitle)
    if not cod and not ent and not resp:
        continue
    
    if grupo not in areas:
        areas[grupo] = []
    
    areas[grupo].append({
        'og_full': og_full,
        'oe_full': oe_full,
        'kpi': kpi,
        'cod': cod,
        'act': act,
        'ent': ent,
        'resp': resp,
        'months': months_active
    })

def months_to_js_filter(months):
    """Convert list of month numbers to JS WEEKS.filter expression"""
    if not months:
        return "[]"
    return f"WEEKS.filter(w=>[{','.join(str(m) for m in months)}].includes(w.month)).map(w=>w.wk)"

def build_seed_js(area_name, rows):
    """Build a JavaScript seed array from rows"""
    # Group by OG -> OE -> acts
    ogs = {}
    for row in rows:
        og_full = row['og_full']
        oe_full = row['oe_full']
        
        # Parse OG code and description
        if '|' in og_full:
            og_parts = og_full.split('|', 1)
            og_code = og_parts[0].strip()
            og_desc = og_parts[1].strip().rstrip('.')
        else:
            og_code = og_full[:5] if og_full else 'OG-XX'
            og_desc = og_full
        
        # Parse OE code and description
        if '|' in oe_full:
            oe_parts = oe_full.split('|', 1)
            oe_code = oe_parts[0].strip()
            oe_desc = oe_parts[1].strip().rstrip('.')
        else:
            oe_code = oe_full[:7] if oe_full else 'OE-XX.X'
            oe_desc = oe_full
        
        og_key = og_code
        oe_key = oe_code
        
        if og_key not in ogs:
            ogs[og_key] = {'code': og_code, 'desc': og_desc, 'oes': {}}
        
        if oe_key not in ogs[og_key]['oes']:
            ogs[og_key]['oes'][oe_key] = {'code': oe_code, 'desc': oe_desc, 'kpi': row['kpi'], 'acts': []}
        
        if row['cod']:  # Only add if there's a code
            ogs[og_key]['oes'][oe_key]['acts'].append({
                'cod': row['cod'],
                'act': row['act'],
                'ent': row['ent'],
                'resp': row['resp'],
                'months': row['months']
            })
    
    # Build JS string
    lines = []
    lines.append(f"const {area_name}_SEED = [")
    
    for og_key in sorted(ogs.keys()):
        og = ogs[og_key]
        og_desc_escaped = og['desc'].replace('"', '\\"').replace("'", "\\'")
        lines.append(f' {{og:"{og["code"]}",og_desc:"{og_desc_escaped}",oes:[')
        
        for oe_key in sorted(og['oes'].keys()):
            oe = og['oes'][oe_key]
            oe_desc_escaped = oe['desc'].replace('"', '\\"').replace("'", "\\'")
            kpi_escaped = oe['kpi'].replace('"', '\\"').replace("'", "\\'")
            lines.append(f'  {{oe:"{oe["code"]}",oe_desc:"{oe_desc_escaped}",kpi:"{kpi_escaped}",acts:[')
            
            for a in oe['acts']:
                act_escaped = a['act'].replace('"', '\\"').replace("'", "\\'")
                ent_escaped = a['ent'].replace('"', '\\"').replace("'", "\\'")
                resp_escaped = a['resp'].replace('"', '\\"').replace("'", "\\'")
                weeks_expr = months_to_js_filter(a['months'])
                lines.append(f'   {{cod:"{a["cod"]}",act:"{act_escaped}",ent:"{ent_escaped}",resp:"{resp_escaped}",weeks:{weeks_expr}}},')
            
            lines.append('  ]},')
        
        lines.append(' ]},')
    
    lines.append('];')
    return '\n'.join(lines)

# Generate for both areas
for area_key, area_name in [('Inmobiliarias', 'INMOB'), ('Constructoras', 'CONST')]:
    if area_key in areas:
        js_code = build_seed_js(area_name, areas[area_key])
        print(f"\n// === {area_key.upper()} ===")
        print(js_code)
        print(f"\n// Total activities for {area_key}: {len(areas[area_key])}")
