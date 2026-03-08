import openpyxl

wb = openpyxl.load_workbook('Plan de Gestión 2026 - Entregable final.xlsx', data_only=True)
ws = wb['PTOTAL']

rows = []
for r in range(2, ws.max_row+1):
    row = {}
    for c in range(1, ws.max_column+1):
        hdr = ws.cell(1, c).value
        val = ws.cell(r, c).value
        if hdr:
            row[str(hdr).strip()] = val
    rows.append(row)

# Group by Grupo
groups = {}
for row in rows:
    g = row.get('Grupo','')
    if g:
        g = str(g).strip()
        if g not in groups:
            groups[g] = []
        groups[g].append(row)

month_cols = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
month_map = {m: i+1 for i, m in enumerate(month_cols)}

def get_months(row):
    months = []
    for m in month_cols:
        v = row.get(m)
        if v is not None and str(v).strip() not in ('', 'nan', 'None'):
            months.append(month_map[m])
    return months

def esc(s):
    if s is None:
        return ''
    s = str(s).strip()
    s = s.replace('\\', '\\\\').replace("'", "\\'").replace('"', '\\"').replace('\n', ' ').replace('\r', '')
    return s

def parse_area(data):
    ogs = []
    cur_og = None
    cur_oe = None
    
    for row in data:
        og_val = str(row.get('OG','')).strip() if row.get('OG') else ''
        oe_val = str(row.get('OE','')).strip() if row.get('OE') else ''
        kpi_val = str(row.get('KPI','')).strip() if row.get('KPI') else ''
        cod_val = str(row.get('Código','')).strip() if row.get('Código') else ''
        act_val = str(row.get('Actividad','')).strip() if row.get('Actividad') else ''
        ent_val = str(row.get('Entregable','')).strip() if row.get('Entregable') else ''
        resp_val = str(row.get('Responsable','')).strip() if row.get('Responsable') else ''
        
        if not og_val and not oe_val and not act_val and not cod_val:
            continue
            
        if og_val.startswith('OG-'):
            og_desc = act_val if act_val and not cod_val else ''
            cur_og = {'og': og_val, 'og_desc': og_desc, 'oes': []}
            ogs.append(cur_og)
            cur_oe = None
            if oe_val.startswith('OE-'):
                cur_oe = {'oe': oe_val, 'oe_desc': act_val if not og_desc else '', 'kpi': kpi_val, 'acts': []}
                if cur_og:
                    cur_og['oes'].append(cur_oe)
            continue
        
        if oe_val.startswith('OE-'):
            oe_desc = act_val if act_val and not cod_val else ''
            cur_oe = {'oe': oe_val, 'oe_desc': oe_desc, 'kpi': kpi_val, 'acts': []}
            if cur_og:
                cur_og['oes'].append(cur_oe)
            continue
        
        if cod_val and act_val and act_val not in ('nan', 'None'):
            months = get_months(row)
            act_obj = {
                'cod': cod_val,
                'act': act_val,
                'ent': ent_val if ent_val not in ('nan', 'None') else '',
                'resp': resp_val if resp_val not in ('nan', 'None') else '',
                'months': months
            }
            if cur_oe:
                cur_oe['acts'].append(act_obj)
    
    return ogs

def render_seed(var_name, ogs):
    lines = []
    lines.append(f'const {var_name} = [')
    for og in ogs:
        lines.append(f' {{og:"{esc(og["og"])}",og_desc:"{esc(og["og_desc"])}",oes:[')
        for oe in og['oes']:
            lines.append(f'  {{oe:"{esc(oe["oe"])}",oe_desc:"{esc(oe["oe_desc"])}",kpi:"{esc(oe["kpi"])}",acts:[')
            for act in oe['acts']:
                ms = act['months']
                if ms:
                    months_filter = ','.join(str(m) for m in ms)
                    weeks_expr = f'WEEKS.filter(w=>[{months_filter}].includes(w.month)).map(w=>w.wk)'
                else:
                    weeks_expr = '[]'
                lines.append(f'   {{cod:"{esc(act["cod"])}",act:"{esc(act["act"])}",ent:"{esc(act["ent"])}",resp:"{esc(act["resp"])}",weeks:{weeks_expr}}},')
            lines.append('  ]},')
        lines.append(' ]},')
    lines.append('];')
    return '\n'.join(lines)

# Write directly to file with UTF-8 encoding
with open('seeds_output_v2.js', 'w', encoding='utf-8') as f:
    for gname, gdata in groups.items():
        f.write(f'\n// === {gname.upper()} ===\n')
        ogs = parse_area(gdata)
        
        if 'inmob' in gname.lower() or 'inmobil' in gname.lower():
            vname = 'INMOB_SEED'
        elif 'construct' in gname.lower():
            vname = 'CONST_SEED'
        else:
            vname = gname.upper().replace(' ','_') + '_SEED'
        
        f.write(render_seed(vname, ogs))
        total = sum(len(a['acts']) for og in ogs for a in og['oes'])
        f.write(f'\n\n// Total activities for {gname}: {total}\n')

print("Done! Written to seeds_output_v2.js")
