import openpyxl
wb = openpyxl.load_workbook('Plan de Gestión 2026 - Entregable final.xlsx', data_only=True)
ws = wb['PTOTAL']

print('Columnas completas (fila 4):')
for c in range(1, ws.max_column+1):
    v = ws.cell(4, c).value
    if v:
        print(f'  C{c}: {v}')

print(f'\nTotal filas: {ws.max_row}, Total cols: {ws.max_column}')

# Ver todas las filas
print('\n--- DATOS ---')
for r in range(4, ws.max_row+1):
    grupo = ws.cell(r, 2).value or ''
    og = ws.cell(r, 3).value or ''
    oe = ws.cell(r, 4).value or ''
    kpi = ws.cell(r, 5).value or ''
    cod = ws.cell(r, 6).value or ''
    act = ws.cell(r, 7).value or ''
    ent = ws.cell(r, 8).value or ''
    resp = ws.cell(r, 9).value or ''
    meses = []
    for c in range(10, 22):
        v = ws.cell(r, c).value
        meses.append('1' if v and str(v).strip() else '0')
    mstr = '|'.join(meses)
    if any([grupo, og, oe, cod, act]):
        g_short = str(grupo)[:15] if grupo else ''
        og_short = str(og)[:60] if og else ''
        oe_short = str(oe)[:70] if oe else ''
        print(f'R{r}: G=[{g_short}] OG=[{og_short}] OE=[{oe_short}] KPI=[{kpi}] Cod=[{cod}] Act=[{act}] Ent=[{ent}] Resp=[{resp}] M=[{mstr}]')
